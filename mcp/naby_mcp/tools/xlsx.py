"""XLSX 생성 — 기존 guild-suro-tracker/scripts/generate_xlsx.py 서식 이식.

산출물: 순위 / 닉네임 / 직업 / 지하수로 (지하수로 내림차순)
TOP3 금·은·동 음영, 헤더 고정, 맑은 고딕, 천단위 콤마.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "맑은 고딕"


def build_xlsx(rows: list[dict], out_path: str) -> dict:
    """rows: [{"name","job","score"}, ...] → xlsx 저장."""
    data = sorted(rows, key=lambda r: -int(r["score"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "지하수로"

    hf = PatternFill("solid", fgColor="2B2B2B")
    hfont = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    gold = PatternFill("solid", fgColor="FFF3C4")
    silver = PatternFill("solid", fgColor="ECECEC")
    bronze = PatternFill("solid", fgColor="F6E0C8")
    thin = Side(style="thin", color="D9D9D9")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal="center", vertical="center")
    L = Alignment(horizontal="left", vertical="center")
    R = Alignment(horizontal="right", vertical="center")

    ws.append(["순위", "닉네임", "직업", "지하수로"])
    for c in ws[1]:
        c.fill, c.font, c.alignment, c.border = hf, hfont, C, bd

    for i, r in enumerate(data, 1):
        ws.append([i, r["name"], r["job"], int(r["score"])])
        row = ws[i + 1]
        row[0].alignment = C
        row[1].alignment = L
        row[2].alignment = L
        row[3].alignment = R
        row[3].number_format = "#,##0"
        for c in row:
            c.font = Font(name=FONT, size=11)
            c.border = bd
        if i == 1:
            for c in row:
                c.fill = gold
                c.font = Font(name=FONT, size=11, bold=True)
        elif i == 2:
            for c in row:
                c.fill = silver
        elif i == 3:
            for c in row:
                c.fill = bronze

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    return {
        "path": str(out),
        "count": len(data),
        "total_score": sum(int(r["score"]) for r in data),
        "top3": [(r["name"], int(r["score"])) for r in data[:3]],
    }
