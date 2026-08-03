#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""나비 길드 지하수로 XLSX 생성기 + 교차검증기.

사용법:
  python generate_xlsx.py --data data.json --out out.xlsx [--last 지난주.xlsx]

data.json 형식:
  {"rows": [["닉네임","직업", 점수], ...]}   # 순위는 정렬 후 자동 부여

동작:
  - 지하수로(점수) 내림차순 정렬 → 순위/닉네임/직업/지하수로 4컬럼 XLSX
  - TOP3 금·은·동 음영, 헤더 고정, '맑은 고딕', 천단위 콤마
  - '워드브레이커'(오표기) 잔존 여부 자동 점검 → 있으면 경고
  - --last 지정 시 지난주 파일과 대조: 신규 / 이탈 / 중복 리포트
의존성: openpyxl  (pip install openpyxl --break-system-packages)
"""
import argparse, json, sys
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "맑은 고딕"


def build(rows, out):
    rows = sorted(rows, key=lambda r: -r[2])
    wb = Workbook(); ws = wb.active; ws.title = "지하수로"
    hf = PatternFill("solid", fgColor="2B2B2B")
    hfont = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    gold = PatternFill("solid", fgColor="FFF3C4")
    silver = PatternFill("solid", fgColor="ECECEC")
    bronze = PatternFill("solid", fgColor="F6E0C8")
    thin = Side(style="thin", color="D9D9D9")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal="center", vertical="center")
    L = Alignment(horizontal="left", vertical="center")
    R = Alignment(horizontal="right", vertical="center")

    ws.append(["순위", "닉네임", "직업", "지하수로"])
    for c in ws[1]:
        c.fill = hf; c.font = hfont; c.alignment = C; c.border = bd

    for i, (nick, job, score) in enumerate(rows, 1):
        ws.append([i, nick, job, int(score)])
        r = ws[i + 1]
        r[0].alignment = C; r[1].alignment = L; r[2].alignment = L
        r[3].alignment = R; r[3].number_format = "#,##0"
        for c in r:
            c.font = Font(name=FONT, size=11); c.border = bd
        if i == 1:
            for c in r: c.fill = gold; c.font = Font(name=FONT, size=11, bold=True)
        elif i == 2:
            for c in r: c.fill = silver
        elif i == 3:
            for c in r: c.fill = bronze

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"
    wb.save(out)
    return rows


def load_names(xlsx):
    ws = load_workbook(xlsx).active
    return [r[1].value for r in ws.iter_rows(min_row=2)]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--last", default=None, help="직전 주차 xlsx (교차검증용)")
    a = ap.parse_args()

    data = json.load(open(a.data, encoding="utf-8"))
    rows = data["rows"]

    names = [r[0] for r in rows]
    dup = [k for k, v in Counter(names).items() if v > 1]
    bad_job = sorted({r[1] for r in rows if r[1] == "워드브레이커"})

    rows = build(rows, a.out)

    total = sum(r[2] for r in rows)
    zero = sum(1 for r in rows if r[2] == 0)
    print(f"총원: {len(rows)} | 0점: {zero} | 합계: {total:,}")
    print("TOP3:", [(rows[i][0], f"{rows[i][2]:,}") for i in range(min(3, len(rows)))])
    print("중복 닉:", dup if dup else "없음")
    if bad_job:
        print("⚠️  '워드브레이커' 오표기 잔존 → '윈드브레이커'로 고칠 것")
    print("저장:", a.out)

    if a.last:
        last = set(load_names(a.last)); cur = set(names)
        print("\n[교차검증 vs", a.last.split("/")[-1], "]")
        print("  신규(직전주 없음):", sorted(cur - last) or "없음")
        print("  이탈(이번주 없음):", sorted(last - cur) or "없음")


if __name__ == "__main__":
    main()
