"""
매주 실행: xlsx -> DynamoDB 업로드
사용법: python upload_to_dynamodb.py guild_members_YYYYMMDD_MMDD.xlsx

업로드 정책:
  - 신규 주차 레코드 생성 시, 직전 주차의 회원별 fine_count / last_fine_week 을 name 기준으로 이월
  - 누적 벌금 횟수는 영구 보존 (회원이 한 번이라도 부과되면 이후 모든 주차에 누적값 표시)
"""

import sys
import unicodedata
import boto3
from decimal import Decimal
from datetime import datetime
from openpyxl import load_workbook

# ── 설정 ──────────────────────────────────────────────────────────
TABLE_NAME = "maple_guild"
AWS_REGION  = "ap-northeast-1"
# ────────────────────────────────────────────────────────────────


def _norm(name: str) -> str:
    """한글 이름 매칭용 NFC 정규화. xlsx와 DB의 인코딩 차이로 dict 매칭이
    실패하는 것을 방지한다 (직전 주는 NFD, 새 주는 NFC 같은 케이스)."""
    return unicodedata.normalize("NFC", (name or "").strip())


def get_previous_fines(table) -> dict[str, dict]:
    """직전 주차의 name → {fine_count, last_fine_week, pending_weeks, left_guild, job} 매핑.

    METADATA가 없거나 직전 주차 데이터가 없으면 빈 dict 반환.
    하나라도 의미 있는 상태가 있는 회원만 포함.
    키는 NFC 정규화된 name. job은 xlsx 누락 시 자리표시 레코드에 재사용한다.
    """
    from boto3.dynamodb.conditions import Key

    meta = table.get_item(Key={"week": "METADATA", "rank": 0}).get("Item")
    if not meta or "latest_week" not in meta:
        return {}

    prev_week = meta["latest_week"]
    resp = table.query(
        KeyConditionExpression=Key("week").eq(prev_week) & Key("rank").gt(0)
    )
    items = resp.get("Items", [])

    mapping: dict[str, dict] = {}
    for i in items:
        name = i.get("name")
        if not name:
            continue
        key = _norm(str(name))
        fine_count = int(i.get("fine_count", 0))
        last_fine_week = i.get("last_fine_week")
        pending_weeks = list(i.get("pending_weeks") or [])
        left_guild = bool(i.get("left_guild"))
        spec_down_from_week = i.get("spec_down_from_week")
        spec_down_prior_count = i.get("spec_down_prior_count")
        fine_exempt_weeks = list(i.get("fine_exempt_weeks") or [])
        weapon_tier = i.get("weapon_tier")   # 무기 마크 — 자주 안 바뀌므로 주차 간 이월
        if (fine_count > 0 or last_fine_week or pending_weeks or left_guild
                or spec_down_from_week or fine_exempt_weeks or weapon_tier):
            mapping[key] = {
                "name":                str(name),  # DB에 저장된 원본 표기 보존
                "job":                 str(i.get("job") or "?"),
                "fine_count":          fine_count,
                "last_fine_week":      last_fine_week,
                "pending_weeks":       pending_weeks,
                "left_guild":          left_guild,
                "spec_down_from_week": spec_down_from_week,
                "spec_down_prior_count": spec_down_prior_count,
                "fine_exempt_weeks":   fine_exempt_weeks,
                "weapon_tier":         weapon_tier,
            }
    return mapping


def upload(file_path: str, week: str = None):
    week = week or datetime.now().strftime("%Y%m%d")

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    dynamodb = boto3.resource("dynamodb", region_name=AWS_REGION)
    table    = dynamodb.Table(TABLE_NAME)

    # 직전 주차 누적 벌금 정보 로드 (name 기준 매칭)
    prev_fines = get_previous_fines(table)
    print(f"직전 주차 누적 벌금 보유 회원 수: {len(prev_fines)}명")

    members = []
    carried_count = 0
    rejoin_count = 0
    seen_keys: set[str] = set()
    max_rank = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rank, name, job, score = row
        if rank is None:
            continue
        name_str = str(name).strip()
        name_key = _norm(name_str)
        rank_int = int(rank)
        max_rank = max(max_rank, rank_int)
        score_value = Decimal(str(score if score else 0))
        score_int = int(score if score else 0)
        item = {
            "week":  week,
            "rank":  rank_int,
            "name":  name_str,
            "job":   str(job).strip(),
            "score": score_value,
        }

        prev = prev_fines.get(name_key)
        if prev:
            if prev["fine_count"] > 0:
                item["fine_count"] = prev["fine_count"]
            if prev["last_fine_week"]:
                item["last_fine_week"] = prev["last_fine_week"]
            if prev.get("spec_down_from_week"):
                item["spec_down_from_week"] = prev["spec_down_from_week"]
            if prev.get("spec_down_prior_count"):
                item["spec_down_prior_count"] = prev["spec_down_prior_count"]
            if prev.get("fine_exempt_weeks"):
                item["fine_exempt_weeks"] = prev["fine_exempt_weeks"]
            if prev.get("weapon_tier"):
                item["weapon_tier"] = prev["weapon_tier"]

            # pending_weeks 이월
            new_pending = list(prev["pending_weeks"])
            # 새 주차에 0점이면 현재 주차도 pending에 추가
            if score_int == 0 and week not in new_pending:
                new_pending.append(week)

            # 재가입 처리: 직전 주차에 left_guild=True 였더라도 새 주차에 등장하면 복귀
            # → 새 레코드에는 left_guild 속성을 추가하지 않음 (False 효과)
            if prev["left_guild"]:
                rejoin_count += 1
                # 재가입 시 미해결 누적은 그대로 두지 않고 깨끗한 상태로 재가입
                # (탈퇴 시 pending이 비워졌고, 새로 미참하면 자동 추가)
                new_pending = [week] if score_int == 0 else []

            if new_pending:
                item["pending_weeks"] = new_pending

            carried_count += 1
        else:
            # 신규 회원: 새 주차 0점이면 pending에 추가
            if score_int == 0:
                item["pending_weeks"] = [week]

        members.append(item)
        seen_keys.add(name_key)

    # 자리표시 레코드: xlsx에 없지만 미해결 pending이 남아있는 회원
    # (탈퇴 안 했는데 단순히 새 주차 export에서 누락된 경우 → 사라짐 방지)
    ghost_count = 0
    ghost_base = max(max_rank + 1000, 9000)
    for name_key, prev in prev_fines.items():
        if name_key in seen_keys:
            continue
        if prev["left_guild"]:
            continue  # 탈퇴 처리된 회원은 이월 안 함
        if not prev["pending_weeks"]:
            continue  # 미해결이 없으면 굳이 자리표시 필요 없음

        ghost_item = {
            "week":          week,
            "rank":          ghost_base + ghost_count,
            "name":          prev["name"],
            "job":           prev["job"],
            "score":         Decimal("0"),
            "pending_weeks": list(prev["pending_weeks"]),
            "is_ghost":      True,
        }
        if prev["fine_count"] > 0:
            ghost_item["fine_count"] = prev["fine_count"]
        if prev["last_fine_week"]:
            ghost_item["last_fine_week"] = prev["last_fine_week"]
        if prev.get("spec_down_from_week"):
            ghost_item["spec_down_from_week"] = prev["spec_down_from_week"]
        if prev.get("spec_down_prior_count"):
            ghost_item["spec_down_prior_count"] = prev["spec_down_prior_count"]
        if prev.get("fine_exempt_weeks"):
            ghost_item["fine_exempt_weeks"] = prev["fine_exempt_weeks"]
        members.append(ghost_item)
        ghost_count += 1
        print(f"  [자리표시 생성] {prev['name']} (pending={prev['pending_weeks']})")

    print(
        f"총 {len(members)}명 DynamoDB 업로드 시작... "
        f"(week={week}, 이월 {carried_count}명, 재가입 {rejoin_count}명, 자리표시 {ghost_count}명)"
    )

    with table.batch_writer() as batch:
        for m in members:
            batch.put_item(Item=m)

    table.put_item(Item={
        "week":        "METADATA",
        "rank":        0,
        "latest_week": week,
    })

    print(f"완료! {len(members)}명 업로드됨 -> week={week}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "guild_members_latest.xlsx"
    week = sys.argv[2] if len(sys.argv) >= 3 else None
    upload(path, week)
